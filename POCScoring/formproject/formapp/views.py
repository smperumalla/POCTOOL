from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from .models import Form, Section, Question, Employee, Response, Subsection, Question, FormAssignment
from django.http import HttpResponse
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.http import JsonResponse
import json
from . import views
from collections import namedtuple
from django.http import JsonResponse
from django.urls import reverse
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from django.http import HttpResponseRedirect
from django.utils.text import slugify
import os
from django.contrib import messages
from django.core.mail import send_mail
from decimal import Decimal

def export_to_excel(request, form_id):
    # create a new workbook
    wb = Workbook()
    # delete the default sheet created
    wb.remove(wb.active)

    form = get_object_or_404(Form, id=form_id)
    forms = [form]

    # populate the workbook
    for form in forms:
        # Create a worksheet for each form
        ws = wb.create_sheet(title=form.title)

        # Fetch all form assignments for the current form
        assignments = FormAssignment.objects.filter(form=form)

        # Create list to store all employees
        employees = []

        for assignment in assignments:
            employees.append(assignment.employee.user.username)

        # Write the employee names as headers in the excel sheet
        ws.append(['', '', '', ''] + employees)  # leave first 4 cells empty for employee names

        # Now fetch all sections in the form
        sections = Section.objects.filter(form=form).order_by('order')

        for section in sections:
            ws.append([section.title])

            # Fetch all subsections in the current section
            subsections = Subsection.objects.filter(section=section).order_by('order')

            for subsection in subsections:
                ws.append([f'    {subsection.title}'])  # 4 space indent for subsections

                # Fetch all questions in the current subsection
                questions = Question.objects.filter(subsection=subsection).order_by('order')

                for question in questions:
                    row = [f'        {question.text}']  # 8 space indent for questions

                    # Append the responses of each employee to the current question
                    scores = []
                    for employee in employees:
                        try:
                            response = Response.objects.get(assignment__employee__user__username=employee, question=question)
                            scores.append(response.score)
                        except Response.DoesNotExist:
                            scores.append('')  # Leave empty if no response

                    ws.append(row + [''] * 3 + scores)  # Add empty cells before the scores

    # create a response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={form.title}.xlsx'
    wb.save(response)

    return response

@require_POST
def employee_delete(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    employee.delete()
    return redirect('admin_employee_management')

def send_assignment_email(form, employee):
    # employee's email will be fetched here. I'm just using a placeholder.
    employee_email = employee.user.email

    send_mail(
        'New Form Assignment',
        f'You have been assigned a new form: {form.title}. Please login to your dashboard to access it.',
        'smperumalla@sonicwall.com',  # This is the sender
        [employee_email],  # This is the receiver
    )

@csrf_exempt
def send_email_view(request, assignment_id):
    if request.method == 'POST':
        try:
            assignment = FormAssignment.objects.get(id=assignment_id)
            send_assignment_email(assignment.form, assignment.employee)
            return JsonResponse({'status': 'success'}, status=200)
        except FormAssignment.DoesNotExist:
            return JsonResponse({'error': 'Assignment not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request'}, status=400)


def delete_form(request, form_id):
    form = Form.objects.get(id=form_id)
    form.delete()
    return redirect('admin_dashboard')


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Now we need to check if the user is an admin or a normal employee.
                try:
                    employee = Employee.objects.get(user=user)
                    if user.is_superuser:  # check if the user is superuser
                        return redirect('admin_dashboard')
                    else:
                        return redirect('employee_dashboard')
                except Employee.DoesNotExist:
                    return HttpResponse("Invalid user", status=400)
            else:
                # Invalid login credentials provided.
                return render(request, 'formapp/employee_login.html', {'error': 'Invalid login credentials.'})
        else:
            # Either username or password not provided.
            return render(request, 'formapp/employee_login.html', {'error': 'Both fields are required.'})
    else:
        return render(request, 'formapp/employee_login.html')




def employee_dashboard(request):
    if request.user.is_authenticated:
        user = request.user
        try:
            employee = Employee.objects.get(user=user)
        except Employee.DoesNotExist:
            return HttpResponse("Employee does not exist.")
        
        form_assignments = FormAssignment.objects.filter(employee=employee, completed=False)
        completed_form_assignments = FormAssignment.objects.filter(employee=employee, completed=True)
        
        return render(request, 'formapp/employee_dashboard.html', {
            'employee_name': user.username,
            'form_assignments': form_assignments,
            'completed_form_assignments': completed_form_assignments
        })
    else:
        return HttpResponse("You are not logged in.")

def admin_dashboard(request):
    # Get a list of all forms
    forms = Form.objects.all()

    # Get a list of all completed form assignments
    completed_form_assignments = FormAssignment.objects.filter(completed=True)

    # Constructing a list of dictionaries, each containing the details of a completed form
    completed_forms = []
    for assignment in completed_form_assignments:
        form_details = {
            'form_title': assignment.form.title,
            'employee_name': assignment.employee.user.username,
            'assignment_id': assignment.id,  # include assignment_id in form details
            # include more details as needed
        }
        completed_forms.append(form_details)

    context = {
        'forms': forms,
        'completed_forms': completed_forms,
    }
    return render(request, 'formapp/admin_dashboard.html', context)



def assignment_detail(request, assignment_id):
    assignment = get_object_or_404(FormAssignment, pk=assignment_id)

    question_responses = []
    for question in Question.objects.filter(subsection__section__form=assignment.form):
        try:
            response = Response.objects.get(assignment=assignment, question=question)
            question_responses.append({
                'question': question,
                'response': response,
            })
        except Response.DoesNotExist:
            continue

    context = {
        "assignment": assignment,
        "question_responses": question_responses,
    }
    return render(request, 'formapp/assignment_detail.html', context)





def form_view(request, form_id):
    form = get_object_or_404(Form, id=form_id)
    sections = Section.objects.filter(form=form)
    subsections = Subsection.objects.filter(section__in=sections)
    questions = Question.objects.filter(subsection__in=subsections)
    return render(request, 'formapp/admin_form_detail.html', {
        'form': form,
        'sections': sections,
        'subsections': subsections,
        'questions': questions,
    })

def form_submit(request, assignment_id):
    assignment = get_object_or_404(FormAssignment, id=assignment_id)
    form = assignment.form
    if request.method == 'POST':
        data = json.loads(request.body)  # Load JSON data from request body
        for section in form.section_set.all():
            for subsection in section.subsection_set.all():  # Iterate over subsections within a section
                for question in subsection.question_set.all():  # Iterate over questions within a subsection
                    score = data.get(str(question.id))
                    if score is not None and 0 <= int(score) <= 5:  # Check that score is between 0 and 5
                        Response.objects.create(
                            assignment=assignment,
                            question=question,
                            score=score
                        )
                    else:
                        return JsonResponse({
                            "error": f"Score for question id {question.id} not provided or out of range (0-5)",
                        }, status=400)
        assignment.completed = True  # Mark the assignment as completed
        assignment.save()

        if request.user.is_staff:
            return redirect(reverse('admin_dashboard'))
        else:
            return redirect(reverse('employee_dashboard'))
    else:  # This part will display the form to submit responses
        sections = form.section_set.all()
        subsections = Subsection.objects.filter(section__in=sections)
        questions = Question.objects.filter(subsection__in=subsections)
        return render(request, 'formapp/form_completion.html', {
            'assignment': assignment,
            'sections': sections,
            'subsections': subsections,  # Pass subsections to template
            'questions': questions,
        })





@csrf_exempt
def form_create(request, form_id=None):
    if request.method == 'POST':
        # Load JSON data from request body
        data = json.loads(request.body)

        form_title = data.get('form_title')
        sections_data = data.get('sections')  # this will be a list of dictionaries

        if form_title and sections_data:
            form = Form.objects.create(title=form_title, creator=request.user, is_draft=True)
            
            for section_data in sections_data:
                section = Section.objects.create(title=section_data['title'], form=form)
                for subsection_data in section_data['subsections']:
                    subsection = Subsection.objects.create(title=subsection_data['title'], section=section)
                    for question_data in subsection_data['questions']:
                        Question.objects.create(text=question_data['text'], subsection=subsection)
            
            return JsonResponse({"status": "success"})
        else:
            return JsonResponse({"status": "error", "message": "Form title and sections are required."}, status=400)
    else:
        # this will be run for GET requests
        form = Form.objects.get(id=form_id) if form_id else None
        sections = Section.objects.filter(form=form) if form else []
        subsections = Subsection.objects.filter(section__in=sections) if sections else []
        questions = Question.objects.filter(subsection__in=subsections) if subsections else []

        return render(request, 'formapp/form_creation.html', {
            'form': form,
            'sections': sections,
            'subsections': subsections,
            'questions': questions
        })


def form_assignment(request, form_id=None):
    if form_id is not None:
        form = get_object_or_404(Form, id=form_id)
        if request.method == 'POST':
            # Check if 'send_email' button was pressed
            if 'send_email' in request.POST:
                employee_id = request.POST['send_email']
                try:
                    employee = Employee.objects.get(id=employee_id)
                    send_assignment_email(form, employee)
                    return redirect('admin_dashboard')
                except Employee.DoesNotExist:
                    return HttpResponse("Invalid employee id provided", status=400)
                
            # If not, process form assignment as before
            employee_ids = request.POST.getlist('employees')
            employees = Employee.objects.filter(id__in=employee_ids)
            if employees:
                form.assigned_employees.add(*employees)
                return redirect('admin_dashboard')
            else:
                return HttpResponse("Invalid employee id(s) provided", status=400)
        else:
            employees = Employee.objects.all()
            return render(request, 'formapp/form_assignment.html', {'form': form, 'employees': employees})
    else:
        forms = Form.objects.all()
        return render(request, 'formapp/form_assignment.html', {'forms': forms})



def form_assignment_list(request, form_id):
    form = get_object_or_404(Form, id=form_id)
    employees = form.assigned_employees.all()  # Fetches all employees assigned this form

    return render(request, 'formapp/form_assignment_list.html', {'form': form, 'employees': employees})

    
@csrf_exempt
def form_detail(request, form_id):
    form = get_object_or_404(Form, id=form_id)
    sections = form.sections.all()
    employees = form.assigned_employees.all()

    if request.method == 'POST':
        data = json.loads(request.body)

        # Update form title
        form.title = data.get('form_title')
        form.save()

        # Update existing sections, subsections, and questions and create new ones
        sections_data = data.get('sections', [])
        for section_data in sections_data:
            section_id = section_data.get('id')
            if section_id:
                # Update existing section
                section = Section.objects.get(id=section_id)
                section.title = section_data['title']
                section.save()

                # Handle subsections
                subsections_data = section_data.get('subsections', [])
                for subsection_data in subsections_data:
                    subsection_id = subsection_data.get('id')
                    if subsection_id:
                        # Update existing subsection
                        subsection = Subsection.objects.get(id=subsection_id)
                        subsection.title = subsection_data['title']
                        subsection.save()

                        # Handle questions
                        questions_data = subsection_data.get('questions', [])
                        for question_data in questions_data:
                            question_id = question_data.get('id')
                            if question_id:
                                # Update existing question
                                question = Question.objects.get(id=question_id)
                                question.text = question_data['text']
                                question.save()
                            else:
                                # Create new question
                                Question.objects.create(text=question_data['text'], subsection=subsection)
                    else:
                        # Create new subsection and associated questions
                        new_subsection = Subsection.objects.create(title=subsection_data['title'], section=section)
                        questions_data = subsection_data.get('questions', [])
                        for question_data in questions_data:
                            Question.objects.create(text=question_data['text'], subsection=new_subsection)
            else:
                # Create new section and associated subsections and questions
                new_section = Section.objects.create(title=section_data['title'], form=form)
                subsections_data = section_data.get('subsections', [])
                for subsection_data in subsections_data:
                    new_subsection = Subsection.objects.create(title=subsection_data['title'], section=new_section)
                    questions_data = subsection_data.get('questions', [])
                    for question_data in questions_data:
                        Question.objects.create(text=question_data['text'], subsection=new_subsection)

        return JsonResponse({"status": "success"})

    else:
        context = {
            'form': form,
            'sections': sections,
            'employees': employees,
        }
        return render(request, 'formapp/form_detail.html', context)


@csrf_exempt
def save_draft(request):
    if request.method == 'POST':
        try:
            form_data = json.loads(request.body)

            form_title = form_data.get('form_title')
            if not form_title:
                return JsonResponse({'status': 'error', 'error': 'Invalid form title.'})

            form, created = Form.objects.get_or_create(title=form_title, creator=request.user)
        
            sections = form_data.get('sections')
            if not sections:
                return JsonResponse({'status': 'error', 'error': 'Invalid or missing sections data.'})

            for section_order, section_data in enumerate(sections):
                section_title = section_data.get('title')
                if not section_title:
                    return JsonResponse({'status': 'error', 'error': 'Invalid section title.'})

                section, created = Section.objects.get_or_create(title=section_title, form=form, order=section_order)
                
                subsections = section_data.get('subsections')
                if not subsections:
                    return JsonResponse({'status': 'error', 'error': 'Invalid or missing subsections data.'})

                for subsection_order, subsection_data in enumerate(subsections):
                    subsection_title = subsection_data.get('title')
                    if not subsection_title:
                        return JsonResponse({'status': 'error', 'error': 'Invalid subsection title.'})

                    subsection, created = Subsection.objects.get_or_create(title=subsection_title, section=section, order=subsection_order)

                    questions = subsection_data.get('questions')
                    if not questions:
                        return JsonResponse({'status': 'error', 'error': 'Invalid or missing questions data.'})

                    for question_order, question_data in enumerate(questions):
                        question_text = question_data.get('text')
                        if not question_text:
                            return JsonResponse({'status': 'error', 'error': 'Invalid question text.'})

                        question, created = Question.objects.get_or_create(text=question_text, subsection=subsection, order=question_order)

            return JsonResponse({'status': 'success', 'url': reverse('admin_dashboard')})

        except Exception as e:
            return JsonResponse({'status': 'error', 'error': str(e), 'url': reverse('admin_dashboard')})
    else:
        return JsonResponse({'status': 'error', 'error': 'Invalid request method.'})


def employee_view(request, employee_id):
    employee = get_object_or_404(Employee, id=employee_id)
    forms = employee.form_set.all()  # Fetches all forms associated with this employee
    return render(request, 'formapp/employee_view.html', {'employee': employee, 'forms': forms})

from django.db import IntegrityError

def employee_manage(request):
    user = request.user
    employee, created = Employee.objects.get_or_create(user=user)

    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        new_username = request.POST.get('username')
        new_password = request.POST.get('password')
        new_email = request.POST.get('email')

        # Check if creating a new employee or updating existing
        if employee_id:
            # Updating existing employee
            employee = get_object_or_404(Employee, id=employee_id)

            if new_username:
                try:
                    employee.user.username = new_username
                    employee.user.full_clean()  # This will check the validation for username.
                except ValidationError as e:
                    # Handle the case when username is not unique
                    # You can render the same page with error message.
                    return render(request, 'formapp/admin_employee_management.html', {'error': e.messages})

            if new_password:
                employee.user.set_password(new_password)  # Hashes the password
            if new_email:
                employee.user.email = new_email

            employee.user.save()

        else:
            # Creating new employee
            try:
                user = User.objects.create_user(username=new_username, password=new_password)
                employee, created = Employee.objects.get_or_create(user=user)
            except ValidationError as e:
                # Handle the case when username is not unique
                # You can render the same page with error message.
                return render(request, 'formapp/admin_employee_management.html', {'error': e.messages})
            except IntegrityError:
                # Handle the case when username already exists
                return render(request, 'formapp/admin_employee_management.html', {'error': ['Username already exists']})

        return redirect('admin_dashboard')

    else:
        employees = Employee.objects.all()
        return render(request, 'formapp/admin_employee_management.html', {'employees': employees})


def form_score_view(request, form_id):
    form = get_object_or_404(Form, id=form_id)
    if request.method == 'GET':
        sections = Section.objects.filter(form=form)
        subsections = Subsection.objects.filter(section__in=sections)
        questions = Question.objects.filter(subsection__in=subsections)
        return render(request, 'formapp/form_score.html', {
            'form': form,
            'sections': sections,
            'subsections': subsections,
            'questions': questions,
        })
    elif request.method == 'POST':
        for question in Question.objects.filter(subsection__in=Subsection.objects.filter(section__in=Section.objects.filter(form=form))):
            score = request.POST.get(str(question.id))
            if score is not None and 0 <= int(score) <= 5:
                response, created = Response.objects.get_or_create(
                    assignment=FormAssignment.objects.get(form=form, employee=request.user.employee),
                    question=question,
                )
                response.score = int(score)
                response.save()
        return redirect('employee_dashboard')

@csrf_exempt
def import_form_from_excel(request, form_id=None):
    if request.method == 'POST':
        if 'file' in request.FILES:
            uploaded_file = request.FILES['file']
            # Use the uploaded file's name as the title of the form
            base_name = os.path.splitext(uploaded_file.name)[0]
            form_title = slugify(base_name, allow_unicode=True)
            form = Form.objects.get(id=form_id) if form_id else Form.objects.create(title=form_title, creator=request.user, is_draft=True)
        else:
            return JsonResponse({"status": "error", "message": "No file provided."}, status=400)

        workbook = load_workbook(uploaded_file)
        sheet = workbook.active  # we assume there's only one sheet

        section_order = 0
        subsection_order = 0
        question_order = 0

        for row in sheet.iter_rows():
            for cell in row:
                cell_value = str(cell.value).strip() if cell.value else None
                if cell_value:
                    if cell.font.b:
                        # Section
                        section_order += 1
                        section = Section(title=cell_value, form=form, order=section_order)
                        section.save()
                        subsection_order = 0  # Reset subsection order when a new section begins
                    elif '.' in cell_value and cell_value.count('.') == 1:
                        # Subsection
                        subsection_order += 1
                        subsection = Subsection(title=cell_value, section=section, order=subsection_order)
                        subsection.save()
                        question_order = 0  # Reset question order when a new subsection begins
                    elif '.' in cell_value and cell_value.count('.') == 2:
                        # Question
                        question_order += 1
                        question = Question(text=cell_value, subsection=subsection, order=question_order)
                        question.save()

        return HttpResponseRedirect(reverse('form_edit', args=[form.id]))
    else:
        return JsonResponse({"status": "error", "message": "This method only supports POST requests."}, status=400)



def form_edit(request, form_id):
    form = get_object_or_404(Form, id=form_id)

    if request.method == 'POST':
        try:
            form.title = request.POST.get('form_name', form.title)
            form.save()

            # Print all POST items
            for key, value in request.POST.items():
                print(f'POST item: {key} = {value}')

            # update existing sections, subsections, questions and weights
            for key, value in request.POST.items():
                parts = key.split('_')
                if key.startswith('section_') and not key.startswith('new_section_'):
                    id = parts[1] if len(parts) == 2 else "_".join(parts[1:])
                    if id.isnumeric():
                        section = get_object_or_404(Section, id=id, form=form)
                        section.title = value
                        weight_key = f'section_weight_{id}'
                        if weight_key in request.POST:
                            weight = Decimal(request.POST.get(weight_key, section.weight))
                            print(f'Updating section {id} weight: {weight}')
                            section.weight = weight
                            section.save()
                elif key.startswith('subsection_') and not key.startswith('new_subsection_'):
                    id = parts[1] if len(parts) == 2 else "_".join(parts[1:])
                    if id.isnumeric():
                        subsection = get_object_or_404(Subsection, id=id, section__form=form)
                        subsection.title = value
                        weight_key = f'subsection_weight_{id}'
                        if weight_key in request.POST:
                            weight = Decimal(request.POST.get(weight_key, subsection.weight))
                            print(f'Updating subsection {id} weight: {weight}')
                            subsection.weight = weight
                            subsection.save()
                elif key.startswith('question_') and not key.startswith('new_question_'):
                    id = parts[1] if len(parts) == 2 else "_".join(parts[1:])
                    if id.isnumeric():
                        question = get_object_or_404(Question, id=id, subsection__section__form=form)
                        question.text = value
                        weight_key = f'question_weight_{id}'
                        if weight_key in request.POST:
                            weight = Decimal(request.POST.get(weight_key, question.weight))
                            print(f'Updating question {id} weight: {weight}')
                            question.weight = weight
                            question.save()

            # create new sections, subsections, questions and weights
            for key, value in request.POST.items():
                if key.startswith('new_section_'):
                    section = Section.objects.create(title=value, form=form)
                    weight = Decimal(request.POST.get('new_section_weight_' + str(section.id), '0.0'))
                    print(f'Creating new section {section.id} with weight: {weight}')
                    section.weight = weight
                    section.save()
                elif key.startswith('new_subsection_'):
                    id = key.split('_')[2]
                    section = get_object_or_404(Section, id=id, form=form)
                    subsection = Subsection.objects.create(title=value, section=section)
                    weight = Decimal(request.POST.get('new_subsection_weight_' + str(subsection.id), '0.0'))
                    print(f'Creating new subsection {subsection.id} with weight: {weight}')
                    subsection.weight = weight
                    subsection.save()
                elif key.startswith('new_question_'):
                    id = key.split('_')[2]
                    subsection = get_object_or_404(Subsection, id=id, section__form=form)
                    question = Question.objects.create(text=value, subsection=subsection)
                    weight = Decimal(request.POST.get('new_question_weight_' + str(question.id), '0.0'))
                    print(f'Creating new question {question.id} with weight: {weight}')
                    question.weight = weight
                    question.save()

            messages.success(request, 'Form updated successfully')
        except Exception as e:
            print(f'Error updating form: {str(e)}')
            messages.error(request, f'Error updating form: {str(e)}')

    sections = form.section_set.all()
    context = {'form': form, 'sections': sections}
    return render(request, 'formapp/form_edit.html', context)
